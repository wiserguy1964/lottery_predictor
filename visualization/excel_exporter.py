"""
Excel exporter for backtest results and recommendations
"""
from pathlib import Path
from typing import Dict, List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import Prediction, Draw
from config import LotteryConfig


class ExcelExporter:
    """Export results to formatted Excel files"""
    
    # Color constants
    HEADER_COLOR = '2C5780'  # Blue
    POOL_COLOR = 'C6E0B4'  # Light green
    JOKER_COLOR = 'FFE599'  # Light yellow
    PROFIT_GREEN = 'A9D08E'
    LOSS_RED = 'FFC7CE'
    BORDER_COLOR = 'BFBFBF'
    
    def __init__(self, lottery_config: LotteryConfig):
        """
        Initialize exporter
        
        Args:
            lottery_config: Lottery configuration
        """
        self.config = lottery_config
    
    def export_backtest_results(
        self,
        summaries: Dict[str, Dict[str, float]],
        rankings: List[tuple],
        filename: str
    ):
        """
        Export backtest results to Excel
        
        Args:
            summaries: Strategy summary metrics
            rankings: List of (strategy_id, score) tuples
            filename: Output filename
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Backtest Results"
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "MULTI-STRATEGY COMPARISON RESULTS"
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.HEADER_COLOR, 
                                      end_color=self.HEADER_COLOR, 
                                      fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Subtitle
        ws.merge_cells('A2:H2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Lottery: {self.config.lottery_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        subtitle_cell.alignment = Alignment(horizontal='center')
        subtitle_cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
        
        # Headers
        headers = ['Rank', 'Strategy', 'Joker Acc', 'Avg Main', 'OE Acc', 'HL Acc', 'Sum Acc', 'Score']
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        row = 5
        for rank, (strategy_id, score) in enumerate(rankings, 1):
            if strategy_id not in summaries:
                continue
            
            metrics = summaries[strategy_id]
            
            data = [
                rank,
                strategy_id,
                f"{metrics.get('joker_accuracy', 0):.2%}",
                f"{metrics.get('avg_main_matches', 0):.3f}",
                f"{metrics.get('oe_accuracy', 0):.2%}",
                f"{metrics.get('hl_accuracy', 0):.2%}",
                f"{metrics.get('sum_accuracy', 0):.2%}",
                f"{score:.1f}"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                if rank == 1:
                    cell.fill = PatternFill(start_color=self.PROFIT_GREEN,
                                           end_color=self.PROFIT_GREEN,
                                           fill_type='solid')
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-size columns
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(filename)
        print(f"Backtest results saved to: {filename}")
    
    def export_recommendation(
        self,
        prediction: Prediction,
        current_draw: Draw,
        filename: str
    ):
        """
        Export prediction/recommendation to Excel
        
        Args:
            prediction: Prediction object
            current_draw: Most recent actual draw
            filename: Output filename
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recommendations"
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"RECOMMENDED NUMBERS FOR NEXT {self.config.lottery_name} DRAW"
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                      end_color=self.HEADER_COLOR,
                                      fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Generation info
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        info_cell = ws.cell(row, 1, "Generation Details")
        info_cell.font = Font(bold=True, size=12)
        info_cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
        
        row += 1
        ws.cell(row, 1, "Generated:").font = Font(bold=True)
        ws.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))
        ws.cell(row, 4, "Strategy:").font = Font(bold=True)
        ws.cell(row, 5, prediction.strategy_name)
        
        row += 1
        ws.cell(row, 1, "Confidence:").font = Font(bold=True)
        ws.cell(row, 2, f"{prediction.confidence_score:.1%}")
        
        # Main numbers section
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws.cell(row, 1, "Recommended Main Numbers")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color=self.POOL_COLOR,
                                       end_color=self.POOL_COLOR,
                                       fill_type='solid')
        
        row += 1
        for i, num in enumerate(prediction.main_numbers, 1):
            col = i
            cell = ws.cell(row, col, num)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 8
        
        # Bonus numbers section
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws.cell(row, 1, "Recommended Bonus Numbers")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color=self.JOKER_COLOR,
                                       end_color=self.JOKER_COLOR,
                                       fill_type='solid')
        
        row += 1
        for i, num in enumerate(prediction.bonus_numbers, 1):
            col = i
            cell = ws.cell(row, col, num)
            cell.font = Font(bold=True, size=14, color='CC0000')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        
        # Pattern predictions
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws.cell(row, 1, "Predicted Patterns")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        row += 1
        patterns = [
            ("Odd/Even:", prediction.predicted_oe),
            ("High/Low:", prediction.predicted_hl),
            ("Sum Bracket:", prediction.predicted_sum_bracket)
        ]
        
        for label, value in patterns:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        wb.save(filename)
        print(f"Recommendations saved to: {filename}")


    def export_recommendation_with_wheel(
        self,
        prediction: Prediction,
        current_draw: Draw,
        filename: str,
        wheel_tickets: list,
        wheel_numbers: list
    ):
        """
        Export prediction with wheel tickets to Excel
        
        Args:
            prediction: Prediction object
            current_draw: Most recent actual draw
            filename: Output filename
            wheel_tickets: List of (main_numbers, bonus) tuples
            wheel_numbers: All numbers in the wheel
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wheel Recommendations"
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"WHEEL SYSTEM FOR {self.config.lottery_name}"
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                      end_color=self.HEADER_COLOR,
                                      fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Generation info
        row = 3
        ws.merge_cells(f'A{row}:H{row}')
        info_cell = ws.cell(row, 1, "Wheel Details")
        info_cell.font = Font(bold=True, size=12)
        info_cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
        
        row += 1
        ws.cell(row, 1, "Generated:").font = Font(bold=True)
        ws.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))
        ws.cell(row, 4, "Strategy:").font = Font(bold=True)
        ws.cell(row, 5, prediction.strategy_name)
        
        row += 1
        ws.cell(row, 1, "Numbers in Wheel:").font = Font(bold=True)
        ws.cell(row, 2, f"{len(wheel_numbers)} numbers")
        ws.cell(row, 4, "Tickets Generated:").font = Font(bold=True)
        ws.cell(row, 5, f"{len(wheel_tickets)} tickets")
        
        # Wheel numbers section
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        section_cell = ws.cell(row, 1, "Numbers in Wheel")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color=self.POOL_COLOR,
                                       end_color=self.POOL_COLOR,
                                       fill_type='solid')
        
        row += 1
        for i, num in enumerate(wheel_numbers):
            col = (i % 8) + 1
            if i > 0 and i % 8 == 0:
                row += 1
            
            cell = ws.cell(row, col, num)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Tickets section
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        section_cell = ws.cell(row, 1, "Wheel Tickets")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
        
        row += 1
        # Header for tickets
        headers = ['#', 'Num 1', 'Num 2', 'Num 3', 'Num 4', 'Num 5', 'Joker', 'Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ticket rows
        cost_per_ticket = 0.50
        for idx, (ticket_main, ticket_bonus) in enumerate(wheel_tickets, 1):
            row += 1
            
            # Ticket number
            cell = ws.cell(row, 1, idx)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Main numbers
            for i, num in enumerate(ticket_main):
                cell = ws.cell(row, i + 2, num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Joker
            # Convert bonus to string if it's a list
            bonus_str = ', '.join(map(str, ticket_bonus)) if isinstance(ticket_bonus, (list, tuple)) else str(ticket_bonus)
            cell = ws.cell(row, 7, bonus_str)
            cell.font = Font(bold=True, color='CC0000')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=self.JOKER_COLOR,
                                   end_color=self.JOKER_COLOR,
                                   fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Cost
            cell = ws.cell(row, 8, f"€{cost_per_ticket:.2f}")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Total row
        row += 1
        ws.cell(row, 1, "TOTAL").font = Font(bold=True)
        ws.cell(row, 7, f"{len(wheel_tickets)} tickets").font = Font(bold=True)
        ws.cell(row, 8, f"€{len(wheel_tickets) * cost_per_ticket:.2f}").font = Font(bold=True)
        
        # Auto-size columns
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        wb.save(filename)
        print(f"Wheel recommendations saved to: {filename}")
