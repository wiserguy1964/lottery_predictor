"""
Configuration module that loads lottery parameters from Excel configuration file.
This allows for easy addition of new lottery types without code changes.

Works with both relative and absolute paths.
Project home: C:\DEV_AREA\lottery_predictor
"""
import os
import sys
from pathlib import Path
from typing import Dict, Any, Optional
from dataclasses import dataclass, field
from datetime import datetime
import openpyxl


@dataclass
class LotteryConfig:
    """Configuration for a specific lottery type"""
    lottery_name: str
    game_id: str
    main_count: int  # Number of main numbers drawn (e.g., 5)
    main_pool: int  # Main pool size (e.g., 45)
    main_play_count: int  # How many main numbers to recommend (e.g., 9)
    bonus_count: int  # Number of bonus numbers drawn (e.g., 1 for Joker)
    bonus_pool: int  # Bonus pool size (e.g., 20)
    bonus_play_count: int  # How many bonus numbers to recommend (e.g., 1)
    min_date: datetime  # First draw date
    api_url_pattern: str  # API endpoint pattern
    is_active: bool  # Is this lottery actively used?
    ticket_cost: float = 1.00  # Cost per ticket in euros
    
    @property
    def total_columns(self) -> int:
        """Total columns in data (main + bonus)"""
        return self.main_count + self.bonus_count
    
    @property
    def main_col_start(self) -> int:
        """Starting column index for main numbers (0-based)"""
        return 0
    
    @property
    def bonus_col_start(self) -> int:
        """Starting column index for bonus numbers (0-based)"""
        return self.main_count


@dataclass
class BacktestConfig:
    """Backtesting and analysis parameters"""
    window_size: int = 90
    step_size: int = 5
    min_tests_required: int = 20
    max_wheel_size: int = 20
    wheel_ticket_size: int = 5
    max_tickets_in_wheel: int = 10
    hot_pool_size: int = 12
    cold_pool_size: int = 12
    hot_pick_count: int = 3
    cold_pick_count: int = 2
    fetch_chunk_days: int = 30
    max_retries: int = 3


def find_project_root() -> Path:
    """
    Find the project root directory by looking for lottery_config.xlsx
    
    Returns:
        Path to project root
    """
    # Start from the directory containing this file
    current = Path(__file__).parent.absolute()
    
    # Check if we're already in the project root (has lottery_config.xlsx)
    if (current / 'lottery_config.xlsx').exists():
        return current
    
    # Check parent directory (in case config.py is in a subdirectory)
    if (current.parent / 'lottery_config.xlsx').exists():
        return current.parent
    
    # Check current working directory
    cwd = Path.cwd()
    if (cwd / 'lottery_config.xlsx').exists():
        return cwd
    
    # For Windows, check if we're in C:\DEV_AREA\lottery_predictor
    if os.name == 'nt':
        dev_area = Path('C:/DEV_AREA/lottery_predictor')
        if dev_area.exists() and (dev_area / 'lottery_config.xlsx').exists():
            return dev_area
    
    # Last resort: return current directory
    return current


class ConfigLoader:
    """Loads configuration from Excel file"""
    
    def __init__(self, config_path: Optional[Path] = None):
        """
        Initialize config loader
        
        Args:
            config_path: Path to lottery_config.xlsx. If None, searches automatically.
        """
        if config_path is None:
            # Find project root
            project_root = find_project_root()
            config_path = project_root / 'lottery_config.xlsx'
        
        self.config_path = Path(config_path)
        
        if not self.config_path.exists():
            raise FileNotFoundError(
                f"Configuration file not found: {self.config_path.absolute()}\n"
                f"Searched in:\n"
                f"  - Script directory: {Path(__file__).parent.absolute()}\n"
                f"  - Parent directory: {Path(__file__).parent.parent.absolute()}\n"
                f"  - Working directory: {Path.cwd()}\n"
                f"  - C:\\DEV_AREA\\lottery_predictor (if Windows)\n"
                f"\nPlease ensure lottery_config.xlsx exists in your project directory."
            )
        
        print(f"📁 Using config file: {self.config_path.absolute()}")
        
        self.lottery_configs: Dict[str, LotteryConfig] = {}
        self.backtest_config: Optional[BacktestConfig] = None
        self._load_configs()
    
    def _load_configs(self):
        """Load all configurations from Excel"""
        wb = openpyxl.load_workbook(self.config_path, data_only=True)
        
        # Load lottery configurations
        if 'LOTTERY_CONFIGS' in wb.sheetnames:
            self._load_lottery_configs(wb['LOTTERY_CONFIGS'])
        else:
            raise ValueError("LOTTERY_CONFIGS sheet not found in configuration file")
        
        # Load backtest parameters
        if 'BACKTEST_PARAMS' in wb.sheetnames:
            self._load_backtest_config(wb['BACKTEST_PARAMS'])
        else:
            # Use defaults if sheet not found
            self.backtest_config = BacktestConfig()
        
        wb.close()
    
    def _load_lottery_configs(self, ws):
        """Load lottery configurations from worksheet"""
        # Headers are in row 1, descriptions in row 2, data starts at row 3
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            # Create dict from row
            row_dict = dict(zip(headers, row))
            
            # Parse boolean
            is_active = str(row_dict['is_active']).upper() == 'TRUE'
            
            # Parse date
            min_date = row_dict['min_date']
            if isinstance(min_date, str):
                min_date = datetime.strptime(min_date, '%Y-%m-%d')
            elif isinstance(min_date, datetime):
                pass  # Already a datetime
            else:
                raise ValueError(f"Invalid date format for {row_dict['lottery_name']}")
            
            lottery_config = LotteryConfig(
                lottery_name=str(row_dict['lottery_name']),
                game_id=str(row_dict['game_id']),
                main_count=int(row_dict['main_count']),
                main_pool=int(row_dict['main_pool']),
                main_play_count=int(row_dict['main_play_count']),
                bonus_count=int(row_dict['bonus_count']),
                bonus_pool=int(row_dict['bonus_pool']),
                bonus_play_count=int(row_dict['bonus_play_count']),
                min_date=min_date,
                api_url_pattern=str(row_dict['api_url_pattern']),
                is_active=is_active,
                ticket_cost=float(row_dict.get('ticket_cost', 1.00))
            )
            
            self.lottery_configs[lottery_config.lottery_name] = lottery_config
    
    def _load_backtest_config(self, ws):
        """Load backtest parameters from worksheet"""
        params = {}
        
        # Headers in row 1, data starts at row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            param_name = str(row[0])
            param_value = row[1]
            
            # Convert to int if it's a number
            if isinstance(param_value, (int, float)):
                param_value = int(param_value)
            
            params[param_name] = param_value
        
        self.backtest_config = BacktestConfig(**params)
    
    def get_lottery_config(self, lottery_name: str) -> LotteryConfig:
        """
        Get configuration for a specific lottery
        
        Args:
            lottery_name: Name of the lottery (e.g., 'OPAP_JOKER')
            
        Returns:
            LotteryConfig object
            
        Raises:
            KeyError: If lottery not found
        """
        if lottery_name not in self.lottery_configs:
            available = list(self.lottery_configs.keys())
            raise KeyError(
                f"Lottery '{lottery_name}' not found in configuration.\n"
                f"Available lotteries: {available}"
            )
        
        return self.lottery_configs[lottery_name]
    
    def get_active_lotteries(self) -> Dict[str, LotteryConfig]:
        """Get all active lottery configurations"""
        return {
            name: config 
            for name, config in self.lottery_configs.items() 
            if config.is_active
        }
    

    def _load_backtest_params(self) -> BacktestConfig:
        """Load backtest parameters from Excel BACKTEST_PARAMS sheet"""
        try:
            import pandas as pd
            
            # Read BACKTEST_PARAMS sheet
            df = pd.read_excel(self.config_path, sheet_name='BACKTEST_PARAMS')
            
            # Convert to dict
            params = {}
            for _, row in df.iterrows():
                param_name = row['parameter_name']
                value = row['value']
                
                # Convert to int
                if isinstance(value, (int, float)):
                    params[param_name] = int(value)
                else:
                    params[param_name] = value
            
            # Create BacktestConfig with Excel values
            config = BacktestConfig(
                window_size=params.get('window_size', 90),
                step_size=params.get('step_size', 5),
                min_tests_required=params.get('min_tests_required', 20),
                max_wheel_size=params.get('max_wheel_size', 20),
                wheel_ticket_size=params.get('wheel_ticket_size', 5),
                max_tickets_in_wheel=params.get('max_tickets_in_wheel', 10),
                hot_pool_size=params.get('hot_pool_size', 12),
                cold_pool_size=params.get('cold_pool_size', 12),
                hot_pick_count=params.get('hot_pick_count', 3),
                cold_pick_count=params.get('cold_pick_count', 2),
                fetch_chunk_days=params.get('fetch_chunk_days', 30),
                max_retries=params.get('max_retries', 3)
            )
            
            print(f"  ✓ Loaded backtest config from Excel:")
            print(f"    - window_size: {config.window_size}")
            print(f"    - step_size: {config.step_size}")
            
            return config
            
        except Exception as e:
            print(f"  ⚠ Could not load BACKTEST_PARAMS from Excel: {e}")
            print(f"  Using default values")
            return BacktestConfig()

    def get_backtest_config(self) -> BacktestConfig:
        """Get backtest configuration"""
        if self.backtest_config is None:
            self.backtest_config = self._load_backtest_params()
        return self.backtest_config
    
    def list_lotteries(self) -> list:
        """List all available lottery names"""
        return list(self.lottery_configs.keys())


# Global config loader instance (lazy loaded)
_config_loader: Optional[ConfigLoader] = None


def get_config_loader(config_path: Optional[Path] = None) -> ConfigLoader:
    """
    Get the global config loader instance
    
    Args:
        config_path: Optional path to config file. Only used on first call.
        
    Returns:
        ConfigLoader instance
    """
    global _config_loader
    
    if _config_loader is None:
        _config_loader = ConfigLoader(config_path)
    
    return _config_loader


def get_lottery_config(lottery_name: str = 'OPAP_JOKER') -> LotteryConfig:
    """
    Convenience function to get lottery configuration
    
    Args:
        lottery_name: Name of lottery (default: OPAP_JOKER)
        
    Returns:
        LotteryConfig object
    """
    loader = get_config_loader()
    return loader.get_lottery_config(lottery_name)


def get_backtest_config() -> BacktestConfig:
    """
    Convenience function to get backtest configuration
    
    Returns:
        BacktestConfig object
    """
    loader = get_config_loader()
    return loader.get_backtest_config()


if __name__ == '__main__':
    # Test configuration loading
    print("=" * 70)
    print("TESTING CONFIGURATION LOADER")
    print("=" * 70)
    
    try:
        print(f"\n🔍 Looking for lottery_config.xlsx...")
        print(f"   Script location: {Path(__file__).parent.absolute()}")
        print(f"   Working directory: {Path.cwd()}")
        print(f"   Project root: {find_project_root()}")
        
        loader = ConfigLoader()
        
        print("\n✅ Configuration loaded successfully!")
        print(f"📁 Config file: {loader.config_path.absolute()}")
        
        print("\n📊 Available lotteries:")
        for name in loader.list_lotteries():
            config = loader.get_lottery_config(name)
            print(f"  - {name}: {config.main_count}/{config.main_pool} + "
                  f"{config.bonus_count}/{config.bonus_pool} "
                  f"(Active: {'✓' if config.is_active else '✗'})")
        
        print("\n⚙️ Backtest configuration:")
        backtest = loader.get_backtest_config()
        print(f"  Window size: {backtest.window_size}")
        print(f"  Step size: {backtest.step_size}")
        print(f"  Hot pool size: {backtest.hot_pool_size}")
        
        # Test getting specific lottery
        print("\n🎯 Testing OPAP Joker config:")
        opap = loader.get_lottery_config('OPAP_JOKER')
        print(f"  Game ID: {opap.game_id}")
        print(f"  Main numbers: {opap.main_count} from {opap.main_pool}")
        print(f"  Bonus numbers: {opap.bonus_count} from {opap.bonus_pool}")
        print(f"  First draw: {opap.min_date}")
        print(f"  Play count: {opap.main_play_count} main, {opap.bonus_play_count} bonus")
        
        print("\n✅ All tests passed!")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()